"""Integration & contract tests for Eel RPC bridge endpoints."""

import os
import tempfile
from unittest.mock import patch

import openpyxl

import src.app.eel_bridge as bridge


def setup_function():
    """Reset global workspace forest and file session before each test."""
    bridge.forest.root_nodes.clear()
    bridge.current_file_path = None


def test_eel_add_and_delete_node():
    res1 = bridge.add_node(None, "RootA", is_container=True)
    assert res1["success"] is True
    root_id = res1["node"]["id"]

    res2 = bridge.add_node(root_id, "ChildA", is_container=False)
    assert res2["success"] is True
    child_id = res2["node"]["id"]

    assert len(res2["roots"]) == 1
    assert len(res2["roots"][0]["children"]) == 1

    # Delete child node
    del_child = bridge.delete_node(child_id)
    assert del_child["success"] is True
    assert len(del_child["roots"][0]["children"]) == 0

    # Delete root node
    del_root = bridge.delete_node(root_id)
    assert del_root["success"] is True
    assert len(del_root["roots"]) == 0


def test_eel_update_node():
    res1 = bridge.add_node(None, "Finance", is_container=True)
    root_id = res1["node"]["id"]
    res2 = bridge.add_node(root_id, "Budget_2026", is_container=False)
    child_id = res2["node"]["id"]

    # 1. Update node name
    update_res = bridge.update_node(root_id, name="Accounting")
    assert update_res["success"] is True
    assert update_res["node"]["name"] == "Accounting"
    assert update_res["roots"][0]["name"] == "Accounting"
    assert update_res["roots"][0]["children"][0]["absolute_path"] == "Accounting\\Budget_2026"

    # 2. Update child node name and data type
    update_child = bridge.update_node(child_id, name="Annual_Budget", data_type="Currency")
    assert update_child["success"] is True
    assert update_child["roots"][0]["children"][0]["name"] == "Annual_Budget"
    assert update_child["roots"][0]["children"][0]["data_type"] == "Currency"
    assert update_child["roots"][0]["children"][0]["absolute_path"] == "Accounting\\Annual_Budget"

    # 3. Reject empty name
    bad_rename = bridge.update_node(root_id, name="   ")
    assert bad_rename["success"] is False
    assert "empty" in bad_rename["error"]


def test_eel_move_node_cycle_rejection():
    res1 = bridge.add_node(None, "Parent", is_container=True)
    p_id = res1["node"]["id"]

    res2 = bridge.add_node(p_id, "Child", is_container=True)
    c_id = res2["node"]["id"]

    # Attempt to move Parent into Child -> Should reject cycle
    move_res = bridge.move_node(p_id, c_id, "NEST_CHILD")
    assert move_res["success"] is False
    assert "descendant" in move_res["rejection_reason"]


def test_eel_sidebar_reorganizer_rpc_flow():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        # Create Excel file with 2 sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "SheetA"
        ws1.cell(row=1, column=1, value=" Header 2 ")
        ws1.cell(row=1, column=2, value="Header 1")

        ws2 = wb.create_sheet(title="SheetB")
        ws2.cell(row=1, column=1, value="Category")
        wb.save(tmp_path)
        wb.close()

        # 1. Test import_excel_file with hierarchical headers & multi-sheet headers dictionary
        res_import = bridge.import_excel_file(tmp_path)
        assert res_import["success"] is True
        assert res_import["sheets"] == ["SheetA", "SheetB"]
        assert res_import["active_sheet"] == "SheetA"
        assert res_import["headers"] == ["Header 2", "Header 1"]  # original column sequence
        assert "all_headers" in res_import
        assert res_import["all_headers"]["SheetA"] == ["Header 2", "Header 1"]
        assert res_import["all_headers"]["SheetB"] == ["Category"]
        assert "roots" in res_import
        assert len(res_import["roots"]) == 2

        # 2. Modify SheetA by adding a node
        add_res_a = bridge.add_node(None, "CustomNodeA", is_container=True)
        assert add_res_a["success"] is True

        # 3. Test switch_active_sheet to SheetB and modify SheetB
        res_switch = bridge.switch_active_sheet("SheetB")
        assert res_switch["success"] is True
        assert res_switch["sheet_name"] == "SheetB"
        assert res_switch["headers"] == ["Category"]
        add_res_b = bridge.add_node(None, "CustomNodeB", is_container=True)
        assert add_res_b["success"] is True

        # 5. Switch back to SheetA and verify CustomNodeA is preserved in memory
        res_back_a = bridge.switch_active_sheet("SheetA")
        assert res_back_a["success"] is True
        root_names_a = [r["name"] for r in res_back_a["roots"]]
        assert "CustomNodeA" in root_names_a

        # 6. Test save_template_sync exporting both modified sheets simultaneously
        res_sync = bridge.save_template_sync(tmp_path)
        assert res_sync["success"] is True
        assert res_sync["template_path"] == tmp_path

        # Verify exported template file has custom changes across both sheets with max_row == 1
        wb_check = openpyxl.load_workbook(tmp_path)
        assert wb_check.sheetnames == ["SheetA", "SheetB"]
        ws_a = wb_check["SheetA"]
        ws_b = wb_check["SheetB"]

        # Collect headers in Row 1 for both sheets
        headers_a = [ws_a.cell(row=1, column=c).value for c in range(1, ws_a.max_column + 1)]
        headers_b = [ws_b.cell(row=1, column=c).value for c in range(1, ws_b.max_column + 1)]

        assert "CustomNodeA" in headers_a
        assert "CustomNodeB" in headers_b
        assert ws_a.max_row == 1
        assert ws_b.max_row == 1
        wb_check.close()
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


@patch("src.hierarchy_lib.services.dialog_service.filedialog.askopenfilename")
@patch("src.hierarchy_lib.services.dialog_service.filedialog.asksaveasfilename")
@patch("src.hierarchy_lib.services.dialog_service.tk.Tk")
def test_eel_file_dialog_rpc_endpoints(mock_tk, mock_asksave, mock_askopen):
    mock_askopen.return_value = "E:/Data/test_import.xlsx"
    mock_asksave.return_value = "E:/Data/Шаблон_test_import.xlsx"

    open_res = bridge.open_file_dialog()
    assert open_res["success"] is True
    assert open_res["cancelled"] is False
    assert open_res["file_path"] == "E:/Data/test_import.xlsx"

    # Set current_file_path to verify default Шаблон_ naming
    bridge.current_file_path = "E:/Data/test_import.xlsx"
    save_res = bridge.save_file_dialog()
    assert save_res["success"] is True
    assert save_res["cancelled"] is False
    assert save_res["file_path"] == "E:/Data/Шаблон_test_import.xlsx"
    mock_asksave.assert_called_with(
        parent=mock_tk(),
        title="Save Reorganized Excel File",
        defaultextension=".xlsx",
        initialfile="Шаблон_test_import.xlsx",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
    )


def test_eel_add_node_with_zone_positioning():
    # 1. Create root node A and child node B
    res_root = bridge.add_node(None, "RootA", is_container=True)
    root_id = res_root["node"]["id"]

    res_child_b = bridge.add_node(root_id, "NodeB", is_container=True)
    b_id = res_child_b["node"]["id"]

    # 2. Add header BEFORE NodeB
    res_before = bridge.add_node(name="HeaderBefore", is_container=False, target_id=b_id, zone="BEFORE_SIBLING")
    assert res_before["success"] is True

    # 3. Add header AFTER NodeB
    res_after = bridge.add_node(name="HeaderAfter", is_container=False, target_id=b_id, zone="AFTER_SIBLING")
    assert res_after["success"] is True

    # 4. Add header NEST_CHILD inside NodeB
    res_inside = bridge.add_node(name="HeaderInside", is_container=False, target_id=b_id, zone="NEST_CHILD")
    assert res_inside["success"] is True

    tree_roots = bridge.forest.to_dict()["roots"]
    root_node = tree_roots[0]
    children_names = [c["name"] for c in root_node["children"]]
    assert children_names == ["HeaderBefore", "NodeB", "HeaderAfter"]

    node_b = root_node["children"][1]
    assert node_b["children"][0]["name"] == "HeaderInside"


def test_eel_hierarchical_excel_import_and_switch():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.cell(row=1, column=1, value=r"Company\Sales\Orders")
        ws1.cell(row=1, column=2, value=r"Company\Sales\Customers")
        ws1.cell(row=1, column=3, value=r"Company\HR\Employees")

        ws2 = wb.create_sheet(title="Warehouse")
        ws2.cell(row=1, column=1, value=r"Inventory\Stock\Items")
        wb.save(tmp_path)
        wb.close()

        # Import file: should auto-generate hierarchy for Sales
        res1 = bridge.import_excel_file(tmp_path)
        assert res1["success"] is True
        roots1 = res1["roots"]
        assert len(roots1) == 1
        assert roots1[0]["name"] == "Company"
        folder_names = [c["name"] for c in roots1[0]["children"]]
        assert "Sales" in folder_names
        assert "HR" in folder_names

        # Switch to Warehouse: should regenerate hierarchy for Warehouse
        res2 = bridge.switch_active_sheet("Warehouse")
        assert res2["success"] is True
        roots2 = res2["roots"]
        assert len(roots2) == 1
        assert roots2[0]["name"] == "Inventory"
        assert roots2[0]["children"][0]["name"] == "Stock"
        assert roots2[0]["children"][0]["children"][0]["name"] == "Items"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_eel_update_node_and_type():
    res1 = bridge.add_node(None, "Revenue", is_container=False, data_type="Text")
    node_id = res1["node"]["id"]
    assert res1["node"]["data_type"] == "Text"

    # Update type using update_node
    res_type = bridge.update_node(node_id, data_type="Currency")
    assert res_type["success"] is True
    assert res_type["node"]["data_type"] == "Currency"
    assert res_type["roots"][0]["data_type"] == "Currency"

    # Update both name and type using update_node
    res_both = bridge.update_node(node_id, name="TotalRevenue", data_type="Decimal")
    assert res_both["success"] is True
    assert res_both["node"]["name"] == "TotalRevenue"
    assert res_both["node"]["data_type"] == "Decimal"

    # Reject invalid data type
    res_bad = bridge.update_node(node_id, data_type="UnknownFormat")
    assert res_bad["success"] is False
    assert "Invalid data type" in res_bad["error"]


def test_eel_import_with_data_type_metadata_and_sync():
    with (
        tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_src,
        tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_out,
    ):
        src_path = tmp_src.name
        out_path = tmp_out.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"
        c1 = ws.cell(row=1, column=1, value="Name")
        c1.number_format = "@"
        c2 = ws.cell(row=1, column=2, value="Salary")
        c2.number_format = '"$"#,##0.00'
        c3 = ws.cell(row=1, column=3, value="HireDate")
        c3.number_format = "yyyy-mm-dd"

        wb.save(src_path)
        wb.close()

        # Import session
        res_import = bridge.import_excel_file(src_path)
        assert res_import["success"] is True
        assert "all_headers_meta" in res_import
        sheet_meta = res_import["all_headers_meta"]["Employees"]
        meta_dict = {item["name"]: item["type"] for item in sheet_meta}
        assert meta_dict["Name"] == "Text"
        assert meta_dict["Salary"] == "Currency"
        assert meta_dict["HireDate"] == "Date"

        # Check that parsed root leaf nodes inherited these types
        roots = res_import["roots"]
        salary_node = next(r for r in roots if r["name"] == "Salary")
        assert salary_node["data_type"] == "Currency"

        # Save template sync
        res_sync = bridge.save_template_sync(out_path)
        assert res_sync["success"] is True

        # Verify exported file number formats
        wb_out = openpyxl.load_workbook(out_path)
        ws_out = wb_out["Employees"]
        assert ws_out.cell(row=1, column=1).number_format == "@"
        assert (
            "$" in ws_out.cell(row=1, column=2).number_format or "#,##0" in ws_out.cell(row=1, column=2).number_format
        )
        assert "yy" in ws_out.cell(row=1, column=3).number_format.lower()
        wb_out.close()
    finally:
        if os.path.exists(src_path):
            os.remove(src_path)
        if os.path.exists(out_path):
            os.remove(out_path)


def test_eel_refresh_excel_session_success():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        # 1. Create initial workbook with 2 sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sales"
        c1 = ws1.cell(row=1, column=1, value="Header1")
        c1.number_format = "@"
        c2 = ws1.cell(row=1, column=2, value="Revenue")
        c2.number_format = '"$"#,##0.00'

        ws2 = wb.create_sheet(title="Inventory")
        ws2.cell(row=1, column=1, value="ItemCode")
        wb.save(tmp_path)
        wb.close()

        # Import initial file
        res_import = bridge.import_excel_file(tmp_path)
        assert res_import["success"] is True
        assert res_import["active_sheet"] == "Sales"
        assert res_import["headers"] == ["Header1", "Revenue"]

        # 2. Modify Excel file externally: add HireDate, rename Header1 -> Region, change Revenue -> Decimal
        wb_mod = openpyxl.Workbook()
        ws1_mod = wb_mod.active
        ws1_mod.title = "Sales"
        c1_mod = ws1_mod.cell(row=1, column=1, value="Region")
        c1_mod.number_format = "@"
        c2_mod = ws1_mod.cell(row=1, column=2, value="Revenue")
        c2_mod.number_format = "0.00"
        c3_mod = ws1_mod.cell(row=1, column=3, value="HireDate")
        c3_mod.number_format = "yyyy-mm-dd"

        ws2_mod = wb_mod.create_sheet(title="Inventory")
        ws2_mod.cell(row=1, column=1, value="ItemCode")
        wb_mod.save(tmp_path)
        wb_mod.close()

        # 3. Call refresh_excel_session
        res_refresh = bridge.refresh_excel_session()
        assert res_refresh["success"] is True
        assert res_refresh["active_sheet"] == "Sales"
        assert res_refresh["headers"] == ["Region", "Revenue", "HireDate"]
        assert res_refresh["all_headers"]["Sales"] == ["Region", "Revenue", "HireDate"]

        sales_meta = {m["name"]: m["type"] for m in res_refresh["all_headers_meta"]["Sales"]}
        assert sales_meta["Region"] == "Text"
        assert sales_meta["Revenue"] == "Decimal"
        assert sales_meta["HireDate"] == "Date"

        # Check roots updated
        roots = res_refresh["roots"]
        names = [r["name"] for r in roots]
        assert "HireDate" in names
        assert "Region" in names
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_eel_refresh_excel_session_active_sheet_fallback():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        # Create workbook with SheetA and SheetB
        wb = openpyxl.Workbook()
        ws_a = wb.active
        ws_a.title = "SheetA"
        ws_a.cell(row=1, column=1, value="ColA")

        ws_b = wb.create_sheet(title="SheetB")
        ws_b.cell(row=1, column=1, value="ColB")
        wb.save(tmp_path)
        wb.close()

        # Import and switch to SheetB
        bridge.import_excel_file(tmp_path)
        bridge.switch_active_sheet("SheetB")
        assert bridge.current_active_sheet == "SheetB"

        # Modify file to remove SheetB (leaving only SheetA and new SheetC)
        wb_mod = openpyxl.Workbook()
        ws_new_a = wb_mod.active
        ws_new_a.title = "SheetA"
        ws_new_a.cell(row=1, column=1, value="ColA")
        ws_new_c = wb_mod.create_sheet(title="SheetC")
        ws_new_c.cell(row=1, column=1, value="ColC")
        wb_mod.save(tmp_path)
        wb_mod.close()

        # Refresh: SheetB was deleted, should fall back to SheetA
        res_refresh = bridge.refresh_excel_session()
        assert res_refresh["success"] is True
        assert res_refresh["active_sheet"] == "SheetA"
        assert res_refresh["sheets"] == ["SheetA", "SheetC"]
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_eel_refresh_excel_session_exceptions():
    # 1. No active file session
    bridge.current_file_path = None
    res_none = bridge.refresh_excel_session()
    assert res_none["success"] is False
    assert "No active Excel session" in res_none["error"]

    # 2. File not found on disk
    bridge.current_file_path = "E:/Data/definitely_missing_file_12345.xlsx"
    res_missing = bridge.refresh_excel_session()
    assert res_missing["success"] is False
    assert "not found" in res_missing["error"].lower()

    # 3. Corrupted / non-excel file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(b"NOT_A_VALID_ZIP_OR_XLSX_HEADER")
        corrupted_path = tmp.name

    try:
        bridge.current_file_path = corrupted_path
        res_corrupt = bridge.refresh_excel_session()
        assert res_corrupt["success"] is False
        assert len(res_corrupt["error"]) > 0
    finally:
        if os.path.exists(corrupted_path):
            os.remove(corrupted_path)


def test_eel_get_and_update_settings():
    # 1. Get settings
    get_res = bridge.get_settings()
    assert get_res["success"] is True
    assert "delimiter" in get_res["settings"]
    assert "default_data_type" in get_res["settings"]

    # 2. Update settings
    upd_res = bridge.update_settings(delimiter="/", default_data_type="Decimal")
    assert upd_res["success"] is True
    assert upd_res["settings"]["delimiter"] == "/"
    assert upd_res["settings"]["default_data_type"] == "Decimal"

    # 3. Reset settings
    reset_res = bridge.reset_settings()
    assert reset_res["success"] is True
    assert reset_res["settings"]["delimiter"] == "\\"
    assert reset_res["settings"]["default_data_type"] == "Text"


def test_eel_update_settings_live_tree_recalculation():
    # Build tree: Root -> Sub -> Item
    bridge.reset_settings()
    r = bridge.add_node(None, "Org", is_container=True)
    root_id = r["node"]["id"]
    s = bridge.add_node(root_id, "Dept", is_container=True)
    sub_id = s["node"]["id"]
    bridge.add_node(sub_id, "Emp", is_container=False)

    # Change delimiter to /
    res = bridge.update_settings(delimiter="/")
    assert res["success"] is True
    root = res["roots"][0]
    sub = root["children"][0]
    emp = sub["children"][0]
    assert emp["absolute_path"] == "Org/Dept/Emp"

    # Reset
    bridge.reset_settings()


def test_eel_update_settings_validation_error():
    res = bridge.update_settings(delimiter="")
    assert res["success"] is False
    assert "error" in res
