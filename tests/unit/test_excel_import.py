"""Unit tests for Excel import adapter reading Row 1 path strings."""

import tempfile
import os
import pytest
import openpyxl
from src.hierarchy_lib.adapters.excel_adapter import ExcelHierarchyAdapter


def test_excel_import_single_sheet_path():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value="Root\\Folder\\Item1")
        wb.save(tmp_path)
        wb.close()

        forest = ExcelHierarchyAdapter.import_from_file(tmp_path)
        assert len(forest.root_nodes) == 1
        root = forest.root_nodes[0]
        assert root.name == "Root"

        paths = forest.get_all_leaf_paths()
        assert len(paths) == 1
        assert paths[0] == "Root\\Folder\\Item1"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_excel_import_multi_sheet_multiple_paths():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        ws1.cell(row=1, column=1, value="Company\\HR\\Employee")

        ws2 = wb.create_sheet(title="S2")
        ws2.cell(row=1, column=1, value="Company\\Engineering\\Developer")

        wb.save(tmp_path)
        wb.close()

        forest = ExcelHierarchyAdapter.import_from_file(tmp_path)
        assert len(forest.root_nodes) == 1
        root = forest.root_nodes[0]
        assert root.name == "Company"

        paths = forest.get_all_leaf_paths()
        assert len(paths) == 2
        assert "Company\\HR\\Employee" in paths
        assert "Company\\Engineering\\Developer" in paths
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
