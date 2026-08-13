"""Unit tests for Excel export adapter writing vertical segment cells."""

import tempfile
import os
import pytest
import openpyxl
from src.hierarchy_lib.models.composite import CompositeNode
from src.hierarchy_lib.models.leaf import LeafNode
from src.hierarchy_lib.services.forest import WorkspaceForest
from src.hierarchy_lib.adapters.excel_adapter import ExcelHierarchyAdapter


def test_excel_export_vertical_path_cells():
    forest = WorkspaceForest()

    root = CompositeNode("Root")
    folder = CompositeNode("Folder")
    item = LeafNode("Item")

    folder.add_child(item)
    root.add_child(folder)
    forest.add_root(root)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        count = ExcelHierarchyAdapter.export_to_file(forest, tmp_path)
        assert count == 1

        wb = openpyxl.load_workbook(tmp_path)
        assert len(wb.worksheets) == 1
        ws = wb.worksheets[0]

        # Verify vertical segment cells (one element per cell, one element per row)
        assert ws.cell(row=1, column=1).value == "Root"
        assert ws.cell(row=2, column=1).value == "Folder"
        assert ws.cell(row=3, column=1).value == "Item"
        wb.close()
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
