"""
Unit tests validating ExcelHierarchyAdapter and PathParser against deterministic sample fixtures.
Verifies header extraction, multi-sheet handling, data type inference, unicode support, and round-trip export.
"""

import os
import tempfile

import openpyxl
import pytest

from src.hierarchy_lib.adapters.excel_adapter import ExcelHierarchyAdapter
from src.hierarchy_lib.services.path_parser import PathParserService
from tests.fixtures.generate_fixtures import FIXTURES_DIR, generate_all_fixtures


@pytest.fixture(scope="session", autouse=True)
def ensure_fixtures():
    """Ensures fixture workbooks exist before running tests."""
    if not FIXTURES_DIR.exists() or not list(FIXTURES_DIR.glob("*.xlsx")):
        generate_all_fixtures(FIXTURES_DIR)


def test_standard_hierarchy_fixture():
    """Validates standard 3-level product hierarchy extraction and structure."""
    fixture_path = FIXTURES_DIR / "standard_hierarchy.xlsx"
    assert fixture_path.exists(), f"Fixture missing: {fixture_path}"

    sheets = ExcelHierarchyAdapter.get_sheet_names(str(fixture_path))
    assert sheets == ["Products"]

    headers = ExcelHierarchyAdapter.read_row1_headers(str(fixture_path), "Products")
    assert headers == ["Category", "Subcategory", "Product", "Price", "InStock"]

    headers_and_types = ExcelHierarchyAdapter.read_row1_headers_and_types(str(fixture_path), "Products")
    header_names = [h for h, _ in headers_and_types]
    assert header_names == headers

    # Construct path hierarchy using backslash delimiter
    mock_paths = [
        r"Electronics\Computers\Laptop Pro 15",
        r"Electronics\Computers\Desktop Workstation",
        r"Electronics\Audio\Wireless Headphones",
        r"Electronics\Audio\Bluetooth Speaker",
        r"Home & Living\Kitchen\Espresso Machine",
        r"Home & Living\Kitchen\High-Power Blender",
        r"Home & Living\Bedroom\Orthopedic Pillow",
    ]
    forest = PathParserService.parse_header_paths(mock_paths, delimiter="\\")
    root_names = [r.name for r in forest.root_nodes]
    assert "Electronics" in root_names
    assert "Home & Living" in root_names

    electronics = next(r for r in forest.root_nodes if r.name == "Electronics")
    sub_names = [c.name for c in electronics.children]
    assert set(sub_names) == {"Computers", "Audio"}


def test_deep_ragged_hierarchy_fixture():
    """Validates deep 5-6 level hierarchy with ragged branch depths and sparse headers."""
    fixture_path = FIXTURES_DIR / "deep_hierarchy_ragged.xlsx"
    assert fixture_path.exists()

    headers = ExcelHierarchyAdapter.read_row1_headers(str(fixture_path), "DeepTree")
    assert headers == ["Enterprise", "Division", "Department", "Team", "Role", "PermissionLevel"]

    deep_paths = [
        r"Acme Corp\Technology\Software Dev\Frontend Team\Lead Engineer",
        r"Acme Corp\Technology\Software Dev\Frontend Team\Junior Developer",
        r"Acme Corp\Technology\Software Dev\Backend Team\Senior Architect",
        r"Acme Corp\Technology\DevOps\Site Reliability Engineer",
        r"Acme Corp\Technology\Chief Technology Officer",
        r"Acme Corp\Finance\Accounting\Payroll\Specialist",
        r"Acme Corp\Finance\Audit\Senior Auditor",
    ]
    forest = PathParserService.parse_header_paths(deep_paths)
    assert len(forest.root_nodes) == 1
    root = forest.root_nodes[0]
    assert root.name == "Acme Corp"

    tech = next(c for c in root.children if c.name == "Technology")
    finance = next(c for c in root.children if c.name == "Finance")
    assert len(tech.children) >= 3
    assert len(finance.children) >= 2


def test_multisheet_retail_fixture():
    """Validates multi-sheet workbook reading and per-sheet schema isolation."""
    fixture_path = FIXTURES_DIR / "multisheet_retail.xlsx"
    assert fixture_path.exists()

    sheets = ExcelHierarchyAdapter.get_sheet_names(str(fixture_path))
    assert sheets == ["Store_East", "Store_West", "Warehouse_Central"]

    east_headers = ExcelHierarchyAdapter.read_row1_headers(str(fixture_path), "Store_East")
    west_headers = ExcelHierarchyAdapter.read_row1_headers(str(fixture_path), "Store_West")
    wh_headers = ExcelHierarchyAdapter.read_row1_headers(str(fixture_path), "Warehouse_Central")

    assert east_headers == ["Department", "Sector", "ItemName", "Barcode", "Price"]
    assert west_headers == ["Division", "Group", "ProductSKU", "StockQuantity"]
    assert wh_headers == ["Zone", "RackNumber", "PalletCode", "WeightKg"]

    # Test round-trip export with multi-sheet template
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_out:
        out_path = tmp_out.name

    try:
        sheets_data = {
            "Store_East": [
                {"path": r"Grocery\Bakery\Sourdough", "type": "Text"},
                {"path": r"Apparel\Menswear\T-Shirt", "type": "Text"},
            ],
            "Store_West": [
                {"path": r"Hardware\Tools\Drill", "type": "Text"},
            ],
            "Warehouse_Central": [
                {"path": r"Zone A\Cold Storage\Rack 1", "type": "Text"},
            ],
        }
        ExcelHierarchyAdapter.export_multi_sheet_template(str(fixture_path), sheets_data, out_path)

        # Verify output file
        out_sheets = ExcelHierarchyAdapter.get_sheet_names(out_path)
        assert out_sheets == ["Store_East", "Store_West", "Warehouse_Central"]

        wb_out = openpyxl.load_workbook(out_path)
        ws_east = wb_out["Store_East"]
        assert ws_east.cell(row=1, column=1).value == r"Grocery\Bakery\Sourdough"
        assert ws_east.cell(row=1, column=2).value == r"Apparel\Menswear\T-Shirt"
        wb_out.close()
    finally:
        if os.path.exists(out_path):
            os.remove(out_path)


def test_cyrillic_and_symbols_fixture():
    """Validates full unicode support (Ukrainian Cyrillic, quotes, apostrophes, symbols)."""
    fixture_path = FIXTURES_DIR / "cyrillic_and_symbols.xlsx"
    assert fixture_path.exists()

    sheets = ExcelHierarchyAdapter.get_sheet_names(str(fixture_path))
    assert sheets == ["Каталог Продукції"]

    headers = ExcelHierarchyAdapter.read_row1_headers(str(fixture_path), "Каталог Продукції")
    assert "Категорія №1" in headers
    assert "Підкатегорія (група / підгрупа)" in headers
    assert "Найменування товару «Бренд / Модель»" in headers
    assert "Ціна, ₴ (грн)" in headers

    # Verify building and exporting Cyrillic hierarchy paths
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_out:
        out_path = tmp_out.name

    try:
        cyrillic_paths = [
            {
                "path": "Побутова техніка «Преміум»\\Кухонні комбайни / Блендери\\Блендер 'Швидкий-Pro' №123",
                "type": "Text",
            },
            {"path": 'Оргтехніка & IT\\Ноутбуки 15.6"\\Ноутбук «Дніпро-Tech» 16GB/512GB', "type": "Text"},
        ]
        sheets_data = {"Каталог Продукції": cyrillic_paths}
        ExcelHierarchyAdapter.export_multi_sheet_template(str(fixture_path), sheets_data, out_path)

        out_headers = ExcelHierarchyAdapter.read_row1_headers(out_path, "Каталог Продукції")
        assert out_headers[0] == "Побутова техніка «Преміум»\\Кухонні комбайни / Блендери\\Блендер 'Швидкий-Pro' №123"
        assert out_headers[1] == 'Оргтехніка & IT\\Ноутбуки 15.6"\\Ноутбук «Дніпро-Tech» 16GB/512GB'
    finally:
        if os.path.exists(out_path):
            os.remove(out_path)


def test_data_types_matrix_fixture():
    """Validates data types matrix fixture and column formatting detection."""
    fixture_path = FIXTURES_DIR / "data_types_matrix.xlsx"
    assert fixture_path.exists()

    headers = ExcelHierarchyAdapter.read_row1_headers(str(fixture_path), "TypesMatrix")
    assert headers == ["Section", "StringCode", "QuantityInt", "PriceFloat", "IsActiveBool", "CreatedDate"]

    headers_and_types = ExcelHierarchyAdapter.read_row1_headers_and_types(
        str(fixture_path), "TypesMatrix", default_data_type="Text"
    )
    assert len(headers_and_types) == 6
    names = [name for name, _ in headers_and_types]
    assert names == headers
