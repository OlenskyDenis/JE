"""Unit tests for ExcelHierarchyAdapter Row 1 header reading, sheet management, and horizontal re-export."""

import tempfile
import os
import openpyxl
from src.hierarchy_lib.adapters.excel_adapter import ExcelHierarchyAdapter


def test_get_sheet_names_and_read_row1_headers():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.cell(row=1, column=1, value=" Region ")
        ws1.cell(row=1, column=2, value="Revenue")
        ws1.cell(row=1, column=3, value=" Region ")  # duplicate
        ws1.cell(row=2, column=1, value="Ignore Row 2 Data")

        ws2 = wb.create_sheet(title="Inventory")
        ws2.cell(row=1, column=1, value="Stock ID")
        ws2.cell(row=1, column=2, value=" Quantity ")

        wb.save(tmp_path)
        wb.close()

        # Test sheet listing
        sheet_names = ExcelHierarchyAdapter.get_sheet_names(tmp_path)
        assert sheet_names == ["Sales", "Inventory"]

        # Test Row 1 header extraction (deduplicated, trimmed, original column sequence)
        sales_headers = ExcelHierarchyAdapter.read_row1_headers(tmp_path, "Sales")
        assert sales_headers == ["Region", "Revenue"]

        inv_headers = ExcelHierarchyAdapter.read_row1_headers(tmp_path, "Inventory")
        assert inv_headers == ["Stock ID", "Quantity"]

    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)



def test_export_multi_sheet_template():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_src, \
         tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_out:
        src_path = tmp_src.name
        out_path = tmp_out.name

    try:
        # Create source workbook with 3 sheets and data rows
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.cell(row=1, column=1, value="OldSales1")
        ws1.cell(row=2, column=1, value="SalesData")

        ws2 = wb.create_sheet(title="Inventory")
        ws2.cell(row=1, column=1, value="OldInv1")
        ws2.cell(row=2, column=1, value="InvData")

        ws3 = wb.create_sheet(title="Reference")
        ws3.cell(row=1, column=1, value="RefHeader1")
        ws3.cell(row=2, column=1, value="RefData")
        wb.save(src_path)
        wb.close()

        # Simultaneously export modified leaf paths for Sales and Inventory, leaving Reference unmodified
        sheet_map = {
            "Sales": ["Sales\\North\\A", "Sales\\South\\B"],
            "Inventory": ["Inv\\Warehouse\\Bin1", "Inv\\Warehouse\\Bin2", "Inv\\Warehouse\\Bin3"]
        }

        total_cols = ExcelHierarchyAdapter.export_multi_sheet_template(
            file_path_or_stream=src_path,
            sheet_leaf_paths_map=sheet_map,
            output_path=out_path
        )

        assert total_cols == 2 + 3 + 1  # Sales(2) + Inventory(3) + Reference(1 streamed)

        wb_out = openpyxl.load_workbook(out_path)
        assert wb_out.sheetnames == ["Sales", "Inventory", "Reference"]

        # Check Sales
        ws_sales = wb_out["Sales"]
        assert ws_sales.cell(row=1, column=1).value == "Sales\\North\\A"
        assert ws_sales.cell(row=1, column=2).value == "Sales\\South\\B"
        assert ws_sales.max_row == 1

        # Check Inventory
        ws_inv = wb_out["Inventory"]
        assert ws_inv.cell(row=1, column=1).value == "Inv\\Warehouse\\Bin1"
        assert ws_inv.cell(row=1, column=2).value == "Inv\\Warehouse\\Bin2"
        assert ws_inv.cell(row=1, column=3).value == "Inv\\Warehouse\\Bin3"
        assert ws_inv.max_row == 1

        # Check Reference (streamed original header, zero data rows)
        ws_ref = wb_out["Reference"]
        assert ws_ref.cell(row=1, column=1).value == "RefHeader1"
        assert ws_ref.max_row == 1

        wb_out.close()
    finally:
        if os.path.exists(src_path):
            os.remove(src_path)
        if os.path.exists(out_path):
            os.remove(out_path)


def test_round_trip_parse_and_export():
    from src.hierarchy_lib.services.path_parser import PathParserService

    initial_headers = [r"Company\HR\Employees", r"Company\HR\Salaries", r"Company\Finance\Invoices"]
    forest = PathParserService.parse_header_paths(initial_headers)
    leaf_paths = forest.get_all_leaf_paths()

    assert leaf_paths == initial_headers


def test_read_row1_headers_read_only_streaming_large_sheet():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LargeSheet"
        # Row 1 headers
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Name")
        ws.cell(row=1, column=3, value="Department")

        # Add 1,000 data rows to simulate large sheet
        for r in range(2, 1002):
            ws.cell(row=r, column=1, value=r)
            ws.cell(row=r, column=2, value=f"User_{r}")
            ws.cell(row=r, column=3, value="Engineering")

        wb.save(tmp_path)
        wb.close()

        headers = ExcelHierarchyAdapter.read_row1_headers(tmp_path, "LargeSheet")
        assert headers == ["ID", "Name", "Department"]
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_read_row1_headers_consecutive_empty_cutoff():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CutoffSheet"
        ws.cell(row=1, column=1, value="ColA")
        ws.cell(row=1, column=2, value="ColB")
        # Columns 3 to 12 are left empty (10 consecutive empty cells)
        # Column 13 has a distant header that should be ignored
        ws.cell(row=1, column=13, value="DistantCol")

        wb.save(tmp_path)
        wb.close()

        headers = ExcelHierarchyAdapter.read_row1_headers(tmp_path, "CutoffSheet")
        assert "DistantCol" not in headers
        assert headers == ["ColA", "ColB"]
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_read_row1_headers_small_gap_allowed():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GapSheet"
        ws.cell(row=1, column=1, value="ColA")
        # Columns 2, 3, 4 empty (3 consecutive empty cells < 10)
        ws.cell(row=1, column=5, value="ColB")

        wb.save(tmp_path)
        wb.close()

        headers = ExcelHierarchyAdapter.read_row1_headers(tmp_path, "GapSheet")
        assert headers == ["ColA", "ColB"]
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_read_row1_headers_whitespace_counts_as_empty():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WhitespaceSheet"
        ws.cell(row=1, column=1, value="Header1")
        # Columns 2..11 are whitespace only
        for c in range(2, 12):
            ws.cell(row=1, column=c, value="   ")
        ws.cell(row=1, column=12, value="Header2")

        wb.save(tmp_path)
        wb.close()

        headers = ExcelHierarchyAdapter.read_row1_headers(tmp_path, "WhitespaceSheet")
        assert "Header2" not in headers
        assert headers == ["Header1"]
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)



def test_export_multi_sheet_template_with_cell_number_formats():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_out:
        out_path = tmp_out.name

    try:
        sheet_map = {
            "Finance": [
                {"path": "Company\\Finance\\Revenue", "type": "Currency"},
                {"path": "Company\\Finance\\Date", "type": "Date"},
                {"path": "Company\\Finance\\Units", "type": "Integer"},
                {"path": "Company\\Finance\\TaxRate", "type": "Percentage"},
                {"path": "Company\\Finance\\Notes", "type": "Text"},
            ]
        }

        ExcelHierarchyAdapter.export_multi_sheet_template(
            file_path_or_stream=None,
            sheet_leaf_paths_map=sheet_map,
            output_path=out_path
        )

        wb_out = openpyxl.load_workbook(out_path)
        ws = wb_out["Finance"]

        # Check Row 1 headers
        assert ws.cell(row=1, column=1).value == "Company\\Finance\\Revenue"
        assert ws.cell(row=1, column=2).value == "Company\\Finance\\Date"
        assert ws.cell(row=1, column=3).value == "Company\\Finance\\Units"
        assert ws.cell(row=1, column=4).value == "Company\\Finance\\TaxRate"
        assert ws.cell(row=1, column=5).value == "Company\\Finance\\Notes"

        # Check openpyxl number_format
        assert "$" in ws.cell(row=1, column=1).number_format or ws.cell(row=1, column=1).number_format == openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE or "#,##0" in ws.cell(row=1, column=1).number_format
        assert "yy" in ws.cell(row=1, column=2).number_format.lower() or "dd" in ws.cell(row=1, column=2).number_format.lower()
        assert ws.cell(row=1, column=3).number_format in ("0", "#,##0")
        assert "%" in ws.cell(row=1, column=4).number_format
        assert ws.cell(row=1, column=5).number_format in ("@", "General")
        wb_out.close()
    finally:
        if os.path.exists(out_path):
            os.remove(out_path)


def test_read_row1_headers_and_types_all_formats():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FormatSheet"

        headers_and_fmts = [
            ("Revenue", '"$"#,##0.00'),
            ("HireDate", "yyyy-mm-dd"),
            ("EventTime", "hh:mm:ss"),
            ("CreatedAt", "yyyy-mm-dd hh:mm:ss"),
            ("TaxRate", "0.00%"),
            ("Quantity", "0"),
            ("UnitPrice", "0.00"),
            ("Notes", "@"),
        ]

        for col_idx, (hdr, fmt) in enumerate(headers_and_fmts, start=1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.number_format = fmt

        wb.save(tmp_path)
        wb.close()

        res = ExcelHierarchyAdapter.read_row1_headers_and_types(tmp_path, "FormatSheet")
        res_dict = dict(res)
        assert res_dict["Revenue"] == "Currency"
        assert res_dict["HireDate"] == "Date"
        assert res_dict["EventTime"] == "Time"
        assert res_dict["CreatedAt"] == "DateTime"
        assert res_dict["TaxRate"] == "Percentage"
        assert res_dict["Quantity"] == "Integer"
        assert res_dict["UnitPrice"] == "Decimal"
        assert res_dict["Notes"] == "Text"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_read_row1_headers_and_types_strictly_max_row_1():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MassiveSheet"

        c1 = ws.cell(row=1, column=1, value="Salary")
        c1.number_format = '"$"#,##0.00'
        c2 = ws.cell(row=1, column=2, value="StartDate")
        c2.number_format = "yyyy-mm-dd"
        c3 = ws.cell(row=1, column=3, value="Description")

        # 10,000 data rows to test zero data row loading
        for r in range(2, 10002):
            ws.cell(row=r, column=1, value=50000 + r)
            ws.cell(row=r, column=2, value="2026-01-01")
            ws.cell(row=r, column=3, value=f"User {r}")

        wb.save(tmp_path)
        wb.close()

        import time
        t0 = time.perf_counter()
        res = ExcelHierarchyAdapter.read_row1_headers_and_types(tmp_path, "MassiveSheet")
        elapsed = time.perf_counter() - t0

        assert elapsed < 0.5  # Sub-second execution strictly on Row 1
        res_dict = dict(res)
        assert res_dict["Salary"] == "Currency"
        assert res_dict["StartDate"] == "Date"
        assert res_dict["Description"] == "Text"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_custom_default_data_type_general_columns():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ConfigTest"

        # General/unformatted column
        ws.cell(row=1, column=1, value="UnspecifiedColumn")
        # Explicit date column
        c2 = ws.cell(row=1, column=2, value="CreatedDate")
        c2.number_format = "yyyy-mm-dd"

        wb.save(tmp_path)
        wb.close()

        # Test with default_data_type="Decimal"
        res_decimal = dict(ExcelHierarchyAdapter.read_row1_headers_and_types(
            tmp_path, "ConfigTest", default_data_type="Decimal"
        ))
        assert res_decimal["UnspecifiedColumn"] == "Decimal"
        assert res_decimal["CreatedDate"] == "Date"  # Explicit type preserved

        # Test with default_data_type="Integer"
        res_int = dict(ExcelHierarchyAdapter.read_row1_headers_and_types(
            tmp_path, "ConfigTest", default_data_type="Integer"
        ))
        assert res_int["UnspecifiedColumn"] == "Integer"
        assert res_int["CreatedDate"] == "Date"  # Explicit type preserved
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)



