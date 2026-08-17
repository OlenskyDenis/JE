"""
Generator script for deterministic Excel test fixtures used across JE unit, integration, and E2E tests.
Creates real .xlsx sample files covering standard and edge-case hierarchy structures.
"""

from pathlib import Path

from openpyxl.workbook import Workbook

FIXTURES_DIR = Path(__file__).resolve().parent / "excel_samples"


def generate_standard_hierarchy(output_dir: Path) -> Path:
    """Generates standard 3-level product hierarchy with columns and data types."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Row 1: Headers
    headers = ["Category", "Subcategory", "Product", "Price", "InStock"]
    ws.append(headers)

    # Data Rows
    rows = [
        ["Electronics", "Computers", "Laptop Pro 15", 1299.99, True],
        ["Electronics", "Computers", "Desktop Workstation", 1899.00, True],
        ["Electronics", "Audio", "Wireless Headphones", 149.50, False],
        ["Electronics", "Audio", "Bluetooth Speaker", 79.99, True],
        ["Home & Living", "Kitchen", "Espresso Machine", 249.00, True],
        ["Home & Living", "Kitchen", "High-Power Blender", 89.90, False],
        ["Home & Living", "Bedroom", "Orthopedic Pillow", 45.00, True],
    ]
    for r in rows:
        ws.append(r)

    file_path = output_dir / "standard_hierarchy.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


def generate_deep_hierarchy_ragged(output_dir: Path) -> Path:
    """Generates a deep 5-level hierarchy with ragged branch depths and sparse cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DeepTree"

    headers = ["Enterprise", "Division", "Department", "Team", "Role", "PermissionLevel"]
    ws.append(headers)

    rows = [
        ["Acme Corp", "Technology", "Software Dev", "Frontend Team", "Lead Engineer", 5],
        ["Acme Corp", "Technology", "Software Dev", "Frontend Team", "Junior Developer", 2],
        ["Acme Corp", "Technology", "Software Dev", "Backend Team", "Senior Architect", 5],
        ["Acme Corp", "Technology", "DevOps", None, "Site Reliability Engineer", 4],
        ["Acme Corp", "Technology", None, None, "Chief Technology Officer", 5],
        ["Acme Corp", "Finance", "Accounting", "Payroll", "Specialist", 3],
        ["Acme Corp", "Finance", "Audit", None, "Senior Auditor", 4],
    ]
    for r in rows:
        ws.append(r)

    file_path = output_dir / "deep_hierarchy_ragged.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


def generate_multisheet_retail(output_dir: Path) -> Path:
    """Generates a workbook with 3 distinct sheets, each having its own hierarchy layout."""
    wb = Workbook()

    # Sheet 1: Store East
    ws1 = wb.active
    ws1.title = "Store_East"
    ws1.append(["Department", "Sector", "ItemName", "Barcode", "Price"])
    ws1.append(["Grocery", "Bakery", "Artisan Sourdough", "BAR-00101", 3.50])
    ws1.append(["Grocery", "Dairy", "Organic Whole Milk", "BAR-00102", 2.20])
    ws1.append(["Apparel", "Menswear", "Classic Cotton T-Shirt", "BAR-00201", 19.99])

    # Sheet 2: Store West
    ws2 = wb.create_sheet(title="Store_West")
    ws2.append(["Division", "Group", "ProductSKU", "StockQuantity"])
    ws2.append(["Hardware", "Tools", "Cordless Drill 18V", 42])
    ws2.append(["Hardware", "Fasteners", "Stainless Steel Screws Box", 350])
    ws2.append(["Garden", "Plants", "Indoor Monstera", 15])

    # Sheet 3: Warehouse Central
    ws3 = wb.create_sheet(title="Warehouse_Central")
    ws3.append(["Zone", "RackNumber", "PalletCode", "WeightKg"])
    ws3.append(["Zone A - Cold Storage", "Rack 01", "PLT-9901", 450.5])
    ws3.append(["Zone A - Cold Storage", "Rack 02", "PLT-9902", 320.0])
    ws3.append(["Zone B - Dry Storage", "Rack 05", "PLT-8801", 850.0])

    file_path = output_dir / "multisheet_retail.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


def generate_cyrillic_and_symbols(output_dir: Path) -> Path:
    """Generates a workbook containing Cyrillic characters, quotes, slashes, numbers, and symbols."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог Продукції"

    headers = [
        "Категорія №1",
        "Підкатегорія (група / підгрупа)",
        "Найменування товару «Бренд / Модель»",
        "Артикул-Код",
        "Ціна, ₴ (грн)",
        "Статус / Примітка",
    ]
    ws.append(headers)

    rows = [
        [
            "Побутова техніка",
            "Кухонні комбайни / Блендери",
            "Блендер «Занурювальний-Pro»",
            "АРТ-001/UA",
            1499.50,
            "В наявності (100%)",
        ],
        [
            "Побутова техніка",
            "Кухонні комбайни / Блендери",
            "Подрібнювач 'Швидкий'",
            "АРТ-002/UA",
            850.00,
            "Очікується 15.09",
        ],
        [
            "Оргтехніка & IT",
            'Ноутбуки 15.6"',
            "Ноутбук «Дніпро-Tech» 16GB/512GB",
            "АРТ-009/IT",
            28900.00,
            "Топ продажів!",
        ],
        [
            "Оргтехніка & IT",
            "Мережеве обладнання (Wi-Fi 6)",
            "Маршрутизатор 2.4/5GHz #AX3000",
            "АРТ-015/NET",
            2150.00,
            "Гарантія 24 міс.",
        ],
    ]
    for r in rows:
        ws.append(r)

    file_path = output_dir / "cyrillic_and_symbols.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


def generate_data_types_matrix(output_dir: Path) -> Path:
    """Generates a workbook with explicit column types (Text, Integer, Float, Boolean, Date)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TypesMatrix"

    headers = ["Section", "StringCode", "QuantityInt", "PriceFloat", "IsActiveBool", "CreatedDate"]
    ws.append(headers)

    rows = [
        ["Alpha", "CODE-A1", 100, 49.95, True, "2026-01-15"],
        ["Alpha", "CODE-A2", 0, 0.00, False, "2026-02-20"],
        ["Beta", "CODE-B1", 15420, 1999.99, True, "2026-05-10"],
        ["Gamma", "CODE-G9", -5, 12.345, False, "2026-08-01"],
    ]
    for r in rows:
        ws.append(r)

    file_path = output_dir / "data_types_matrix.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


def generate_all_fixtures(target_dir: Path | None = None) -> list[Path]:
    """Generates all 5 standard test fixtures and returns list of created paths."""
    out_dir = target_dir or FIXTURES_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    created = [
        generate_standard_hierarchy(out_dir),
        generate_deep_hierarchy_ragged(out_dir),
        generate_multisheet_retail(out_dir),
        generate_cyrillic_and_symbols(out_dir),
        generate_data_types_matrix(out_dir),
    ]
    return created


if __name__ == "__main__":
    files = generate_all_fixtures()
    print(f"Successfully generated {len(files)} test fixtures in: {FIXTURES_DIR}")
    for f in files:
        print(f" - {f.name} ({f.stat().st_size} bytes)")
