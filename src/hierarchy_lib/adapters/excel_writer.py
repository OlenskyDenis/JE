"""ExcelWriter using openpyxl for clean multi-sheet template export workbook construction."""

import os
from typing import Any, BinaryIO, Dict, Union

import openpyxl

from src.hierarchy_lib.adapters.excel_reader import ExcelReader


class ExcelWriter:
    """Specialized writer for constructing fresh Excel templates with Row 1 headers."""

    EXCEL_TYPE_FORMAT_MAP = {
        "Text": "@",
        "Integer": "0",
        "Decimal": "0.00",
        "Currency": '"$"#,##0.00',
        "Percentage": "0.00%",
        "Date": "yyyy-mm-dd",
        "Time": "hh:mm:ss",
        "DateTime": "yyyy-mm-dd hh:mm:ss",
        "Boolean": "General",
    }

    @staticmethod
    def export_multi_sheet_template(
        file_path_or_stream: Union[str, BinaryIO, None],
        sheet_leaf_paths_map: Dict[str, Any],
        output_path: str,
    ) -> int:
        """Constructs an openpyxl.Workbook with Row 1 headers and data type formats."""
        new_wb = openpyxl.Workbook()
        has_source = isinstance(file_path_or_stream, str) and os.path.exists(file_path_or_stream)

        if has_source:
            sheet_names = ExcelReader.get_sheet_names(file_path_or_stream)
            for sname in sheet_leaf_paths_map.keys():
                if sname and sname not in sheet_names:
                    sheet_names.append(sname)
        else:
            sheet_names = list(sheet_leaf_paths_map.keys()) if sheet_leaf_paths_map else ["Sheet1"]

        total_cols = 0
        for idx, sname in enumerate(sheet_names):
            if idx == 0:
                ws = new_wb.active
                ws.title = sname
            else:
                ws = new_wb.create_sheet(title=sname)

            if sname in sheet_leaf_paths_map:
                items = sheet_leaf_paths_map[sname]
                for col_idx, item in enumerate(items, start=1):
                    path_str = ""
                    data_type = "Text"

                    if isinstance(item, str):
                        path_str = item
                    elif isinstance(item, dict):
                        path_str = item.get("path") or item.get("name") or ""
                        data_type = item.get("type") or item.get("data_type") or "Text"
                    elif isinstance(item, (tuple, list)):
                        path_str = item[0] if len(item) > 0 else ""
                        data_type = item[1] if len(item) > 1 else "Text"

                    cell = ws.cell(row=1, column=col_idx, value=path_str)
                    fmt = ExcelWriter.EXCEL_TYPE_FORMAT_MAP.get(data_type, "@")
                    cell.number_format = fmt

                total_cols += len(items)
            elif has_source:
                other_headers = ExcelReader.read_row1_headers(file_path_or_stream, sname)
                for col_idx, h_str in enumerate(other_headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=h_str)
                    cell.number_format = "@"
                total_cols += len(other_headers)

        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        new_wb.save(output_path)
        new_wb.close()
        return total_cols
