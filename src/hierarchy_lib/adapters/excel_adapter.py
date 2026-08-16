"""ExcelHierarchyAdapter using openpyxl for self-contained Excel processing."""

import os
from typing import Union, BinaryIO, List, Dict, Any, Tuple, Optional
import openpyxl
from src.hierarchy_lib.services.header_service import HeaderService


class ExcelHierarchyAdapter:
    """Adapter for importing and exporting Composite hierarchy trees to/from Excel (.xlsx) files."""

    EXCEL_TYPE_FORMAT_MAP = {
        "Text": "@",
        "Integer": "0",
        "Decimal": "0.00",
        "Currency": '"$"#,##0.00',
        "Percentage": "0.00%",
        "Date": "yyyy-mm-dd",
        "Time": "hh:mm:ss",
        "DateTime": "yyyy-mm-dd hh:mm:ss",
        "Boolean": "General",
    }

    @staticmethod
    def get_sheet_names(file_path_or_stream: Union[str, BinaryIO]) -> List[str]:
        """Returns the list of worksheet names available in the workbook."""
        wb = openpyxl.load_workbook(file_path_or_stream, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names

    @staticmethod
    def _map_format_to_data_type(
        number_format: Optional[str],
        data_type_flag: Optional[str] = None,
        header_name: Optional[str] = None,
        default_data_type: Optional[str] = None
    ) -> str:
        """Maps Excel number format string and cell data_type flag to one of the 9 standard Excel types."""
        fallback_type = default_data_type if default_data_type is not None else "Text"

        if data_type_flag == "b":
            return "Boolean"

        num_fmt = (number_format or "").strip().lower()
        if not num_fmt or num_fmt in ("@", "general"):
            if data_type_flag == "d":
                return "Date"
            return fallback_type

        has_time = any(t in num_fmt for t in ("h", "s", "am/pm"))
        has_date = any(d in num_fmt for d in ("y", "d", "mmm", "m/d")) or ("m" in num_fmt and not has_time and "/" in num_fmt)

        if "%" in num_fmt:
            return "Percentage"

        if any(curr in num_fmt for curr in ("$", "€", "£", "грн", "₽", "¥", "руб", "¤", "[$")):
            return "Currency"

        if has_date and has_time:
            return "DateTime"
        if has_time and not has_date:
            return "Time"
        if has_date:
            return "Date"

        if "0.00" in num_fmt or "#.00" in num_fmt or "0.0" in num_fmt:
            return "Decimal"

        if num_fmt in ("0", "#,##0", "0_"):
            return "Integer"

        return fallback_type

    @staticmethod
    def read_row1_headers_and_types(
        file_path_or_stream: Union[str, BinaryIO],
        sheet_name: str,
        max_empty_consecutive: int = 10,
        default_data_type: Optional[str] = None
    ) -> List[Tuple[str, str]]:
        """
        Reads Row 1 cells strictly in streaming mode (max_row=1).
        For each valid header, extracts the header name and determines the configured
        Excel column data type based on cell.number_format, column_dimensions.number_format,
        and cell.data_type.
        Stops scanning if max_empty_consecutive empty/None cells are encountered.
        Returns a list of tuples: [(header_name, data_type), ...]
        """
        wb = openpyxl.load_workbook(file_path_or_stream, data_only=False)
        try:
            if sheet_name not in wb.sheetnames:
                return []

            sheet = wb[sheet_name]
            raw_items = []
            consecutive_empty = 0
            max_col = sheet.max_column or 100

            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col_idx)
                val = cell.value
                if val is not None and str(val).strip() != "":
                    consecutive_empty = 0
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    col_dim_fmt = ""
                    if col_letter in sheet.column_dimensions:
                        col_dim_fmt = sheet.column_dimensions[col_letter].number_format or ""

                    num_fmt = cell.number_format or col_dim_fmt or ""
                    detected_type = ExcelHierarchyAdapter._map_format_to_data_type(
                        number_format=num_fmt,
                        data_type_flag=cell.data_type,
                        header_name=str(val).strip(),
                        default_data_type=default_data_type
                    )
                    raw_items.append((str(val).strip(), detected_type))
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= max_empty_consecutive:
                        break

            # Deduplicate preserving original sequence
            seen = set()
            result = []
            for name, dtype in raw_items:
                if name not in seen:
                    seen.add(name)
                    result.append((name, dtype))

            return result
        finally:
            wb.close()

    @staticmethod
    def read_row1_headers(
        file_path_or_stream: Union[str, BinaryIO],
        sheet_name: str,
        max_empty_consecutive: int = 10,
        default_data_type: Optional[str] = None
    ) -> List[str]:
        """
        Reads Row 1 cells exclusively from the specified sheet in read_only streaming mode.
        Ignores rows 2+ completely, and stops scanning if max_empty_consecutive empty/None
        cells are encountered in Row 1.
        Returns clean, deduplicated, sorted headers.
        """
        pairs = ExcelHierarchyAdapter.read_row1_headers_and_types(
            file_path_or_stream,
            sheet_name,
            max_empty_consecutive=max_empty_consecutive,
            default_data_type=default_data_type
        )
        return HeaderService.process_headers([name for name, _ in pairs])

    @staticmethod
    def export_multi_sheet_template(
        file_path_or_stream: Union[str, BinaryIO, None],
        sheet_leaf_paths_map: Dict[str, Any],
        output_path: str
    ) -> int:
        """
        Constructs a fresh openpyxl.Workbook() from scratch containing all original sheets with Row 1 headers only.
        For each sheet in sheet_leaf_paths_map: writes custom leaf_paths across Row 1 columns (A1, B1, C1...) and applies number_format.
        For all other sheets: streams original Row 1 headers without reading data rows.
        Guarantees zero data rows in Row 2+ across all sheets with minimal memory/CPU overhead.
        Returns total columns written across all sheets.
        """
        new_wb = openpyxl.Workbook()

        has_source = isinstance(file_path_or_stream, str) and os.path.exists(file_path_or_stream)

        if has_source:
            sheet_names = ExcelHierarchyAdapter.get_sheet_names(file_path_or_stream)
            for sname in sheet_leaf_paths_map.keys():
                if sname and sname not in sheet_names:
                    sheet_names.append(sname)
        else:
            sheet_names = list(sheet_leaf_paths_map.keys()) if sheet_leaf_paths_map else ["Sheet1"]

        total_cols = 0
        for idx, sname in enumerate(sheet_names):
            if idx == 0:
                ws = new_wb.active
                ws.title = sname
            else:
                ws = new_wb.create_sheet(title=sname)

            if sname in sheet_leaf_paths_map:
                # Target reorganized sheet: write leaf_paths and format types into Row 1
                items = sheet_leaf_paths_map[sname]
                for col_idx, item in enumerate(items, start=1):
                    path_str = ""
                    data_type = "Text"

                    if isinstance(item, str):
                        path_str = item
                    elif isinstance(item, dict):
                        path_str = item.get("path") or item.get("name") or ""
                        data_type = item.get("type") or item.get("data_type") or "Text"
                    elif isinstance(item, (tuple, list)):
                        path_str = item[0] if len(item) > 0 else ""
                        data_type = item[1] if len(item) > 1 else "Text"

                    cell = ws.cell(row=1, column=col_idx, value=path_str)
                    fmt = ExcelHierarchyAdapter.EXCEL_TYPE_FORMAT_MAP.get(data_type, "@")
                    cell.number_format = fmt

                total_cols += len(items)
            elif has_source:
                # Other sheets: stream original Row 1 headers without reading data rows
                other_headers = ExcelHierarchyAdapter.read_row1_headers(file_path_or_stream, sname)
                for col_idx, h_str in enumerate(other_headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=h_str)
                    cell.number_format = "@"
                total_cols += len(other_headers)

        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        new_wb.save(output_path)
        new_wb.close()
        return total_cols
