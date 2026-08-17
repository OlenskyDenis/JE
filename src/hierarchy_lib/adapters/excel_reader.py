"""ExcelReader using openpyxl for streaming read-only Excel header inspection."""

from typing import BinaryIO, List, Optional, Tuple, Union

import openpyxl

from src.hierarchy_lib.services.header_service import HeaderService


class ExcelReader:
    """Specialized reader for streaming row 1 headers and number formats from Excel files."""

    @staticmethod
    def get_sheet_names(file_path_or_stream: Union[str, BinaryIO]) -> List[str]:
        """Returns the list of worksheet names available in the workbook."""
        wb = openpyxl.load_workbook(file_path_or_stream, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names

    @staticmethod
    def map_format_to_data_type(
        number_format: Optional[str],
        data_type_flag: Optional[str] = None,
        header_name: Optional[str] = None,
        default_data_type: Optional[str] = None,
    ) -> str:
        """Maps Excel number format string and cell data_type flag to one of the 9 standard Excel types."""
        fallback_type = default_data_type if default_data_type is not None else "Text"

        if data_type_flag == "b":
            return "Boolean"

        num_fmt = (number_format or "").strip().lower()
        if not num_fmt or num_fmt in ("@", "general"):
            if data_type_flag == "d":
                return "Date"
            return fallback_type

        has_time = any(t in num_fmt for t in ("h", "s", "am/pm"))
        has_date = any(d in num_fmt for d in ("y", "d", "mmm", "m/d")) or (
            "m" in num_fmt and not has_time and "/" in num_fmt
        )

        if "%" in num_fmt:
            return "Percentage"

        if any(curr in num_fmt for curr in ("$", "€", "£", "грн", "₽", "¥", "руб", "¤", "[$")):
            return "Currency"

        if has_date and has_time:
            return "DateTime"
        if has_time and not has_date:
            return "Time"
        if has_date:
            return "Date"

        if "0.00" in num_fmt or "#.00" in num_fmt or "0.0" in num_fmt:
            return "Decimal"

        if num_fmt in ("0", "#,##0", "0_"):
            return "Integer"

        return fallback_type

    @staticmethod
    def read_row1_headers_and_types(
        file_path_or_stream: Union[str, BinaryIO],
        sheet_name: str,
        max_empty_consecutive: int = 10,
        default_data_type: Optional[str] = None,
    ) -> List[Tuple[str, str]]:
        """Reads Row 1 cells strictly in streaming mode and detects column data types."""
        wb = openpyxl.load_workbook(file_path_or_stream, data_only=False)
        try:
            if sheet_name not in wb.sheetnames:
                return []

            sheet = wb[sheet_name]
            raw_items = []
            consecutive_empty = 0
            max_col = sheet.max_column or 100

            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col_idx)
                val = cell.value
                if val is not None and str(val).strip() != "":
                    consecutive_empty = 0
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    col_dim_fmt = ""
                    if col_letter in sheet.column_dimensions:
                        col_dim_fmt = sheet.column_dimensions[col_letter].number_format or ""

                    num_fmt = cell.number_format or col_dim_fmt or ""
                    detected_type = ExcelReader.map_format_to_data_type(
                        number_format=num_fmt,
                        data_type_flag=cell.data_type,
                        header_name=str(val).strip(),
                        default_data_type=default_data_type,
                    )
                    raw_items.append((str(val).strip(), detected_type))
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= max_empty_consecutive:
                        break

            # Deduplicate preserving original sequence
            seen = set()
            result = []
            for name, dtype in raw_items:
                if name not in seen:
                    seen.add(name)
                    result.append((name, dtype))

            return result
        finally:
            wb.close()

    @staticmethod
    def read_row1_headers(
        file_path_or_stream: Union[str, BinaryIO],
        sheet_name: str,
        max_empty_consecutive: int = 10,
        default_data_type: Optional[str] = None,
    ) -> List[str]:
        """Reads clean, deduplicated, sorted Row 1 headers from the specified sheet."""
        pairs = ExcelReader.read_row1_headers_and_types(
            file_path_or_stream,
            sheet_name,
            max_empty_consecutive=max_empty_consecutive,
            default_data_type=default_data_type,
        )
        return HeaderService.process_headers([name for name, _ in pairs])
